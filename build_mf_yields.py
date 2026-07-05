"""
Build mf_yields.csv from MF fortnightly debt-scheme portfolio disclosures.

Input: downloaded disclosure workbooks from 8 AMCs (see FILES below).
Output: mf_yields.csv with columns: isin, yield, as_of, holders, source

Merge logic (additive, AMC priority order):
  - Yield + source come from the FIRST AMC (in priority order) that discloses the ISIN.
  - Every AMC that holds the ISIN is appended to holders.

Yield normalization: sheets that store yield as a decimal fraction
(e.g. 0.073 = 7.3%) are detected via the sheet median and scaled x100.
"""
import re
import statistics
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ISIN_RE = re.compile(r"^IN[A-Z0-9]{10}$")
YIELD_HDR_RE = re.compile(r"^\s*(yield|ytm)\b(?!.*call)", re.I)
YTC_RE = re.compile(r"call|ytc", re.I)
ISIN_HDR_RE = re.compile(r"^\s*isin", re.I)
RATING_HDR_RE = re.compile(r"rating", re.I)
NAME_HDR_RE = re.compile(r"name of|company/issuer|instrument", re.I)

SKIP_SHEETS = re.compile(r"^(index|derivative|deriv)", re.I)


def find_header(cells):
    """Return (isin_col, yield_col, rating_col, name_col) or None."""
    isin_col = yield_col = rating_col = name_col = None
    for j, c in enumerate(cells):
        if c is None:
            continue
        s = str(c).strip()
        if isin_col is None and ISIN_HDR_RE.match(s):
            isin_col = j
        if yield_col is None and YIELD_HDR_RE.match(s) and not YTC_RE.search(s):
            yield_col = j
        if rating_col is None and RATING_HDR_RE.search(s):
            rating_col = j
        if name_col is None and NAME_HDR_RE.search(s) and not ISIN_HDR_RE.match(s):
            name_col = j
    if isin_col is not None and yield_col is not None:
        return isin_col, yield_col, rating_col, name_col
    return None


def parse_yield(v):
    if v is None:
        return None
    s = str(v).strip().replace("%", "").replace(",", "")
    if not s or s in ("-", "NA", "N.A.", "NIL"):
        return None
    try:
        y = float(s)
    except ValueError:
        return None
    if y <= 0:
        return None
    return y


def parse_rows(rows_iter):
    """Yield (isin, yield_raw, rating, name) from a sheet; header re-detected
    whenever encountered (handles UTI's one-big-sheet layout)."""
    cols = None
    for row in rows_iter:
        hdr = find_header(row)
        if hdr:
            cols = hdr
            continue
        if cols is None:
            continue
        ic, yc, rc, nc = cols
        if ic >= len(row):
            continue
        isin = row[ic]
        if isin is None or not ISIN_RE.match(str(isin).strip()):
            continue
        y = parse_yield(row[yc]) if yc < len(row) else None
        rating = str(row[rc]).strip() if rc is not None and rc < len(row) and row[rc] else ""
        name = str(row[nc]).strip() if nc is not None and nc < len(row) and row[nc] else ""
        yield str(isin).strip(), y, rating, name


def normalize_sheet_yields(records):
    """records: list of (isin,y,rating,name). Scale x100 if fraction-style."""
    ys = [r[1] for r in records if r[1] is not None]
    if not ys:
        return records
    med = statistics.median(ys)
    if med < 1:  # stored as decimal fraction
        records = [(i, (y * 100 if y is not None and y < 1.5 else y), r, n)
                   for i, y, r, n in records]
    out = []
    for i, y, r, n in records:
        if y is not None and not (0 < y <= 60):
            y = None
        out.append((i, y, r, n))
    return out


def parse_xlsx(path):
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        if SKIP_SHEETS.match(sn):
            continue
        recs = list(parse_rows(wb[sn].iter_rows(values_only=True)))
        out.extend(normalize_sheet_yields(recs))
    wb.close()
    return out


def parse_xls(path):
    out = []
    xl = pd.ExcelFile(path)
    for sn in xl.sheet_names:
        if SKIP_SHEETS.match(sn):
            continue
        df = xl.parse(sn, header=None)
        rows = (tuple(None if pd.isna(v) else v for v in row)
                for row in df.itertuples(index=False, name=None))
        recs = list(parse_rows(rows))
        out.extend(normalize_sheet_yields(recs))
    return out


def parse_any(path):
    p = str(path)
    return parse_xls(p) if p.lower().endswith(".xls") else parse_xlsx(p)


def build(files_by_amc, as_of_by_amc, out_csv):
    """files_by_amc: ordered dict {amc_name: [paths]} in priority order."""
    merged = {}   # isin -> dict
    for amc, paths in files_by_amc.items():
        seen_this_amc = set()
        for path in paths:
            try:
                recs = parse_any(path)
            except Exception as e:
                print(f"WARN: {amc}: failed to parse {path}: {e}", file=sys.stderr)
                continue
            for isin, y, rating, name in recs:
                entry = merged.get(isin)
                if entry is None:
                    entry = merged[isin] = {
                        "isin": isin, "yield": None, "as_of": None,
                        "holders": [], "source": None,
                        "mf_rating": rating, "name": name,
                    }
                if amc not in entry["holders"] and isin not in seen_this_amc:
                    entry["holders"].append(amc)
                    seen_this_amc.add(isin)
                elif amc not in entry["holders"]:
                    entry["holders"].append(amc)
                if entry["yield"] is None and y is not None:
                    entry["yield"] = round(y, 4)
                    entry["as_of"] = as_of_by_amc[amc]
                    entry["source"] = amc
        print(f"{amc}: cumulative ISINs {len(merged)}")
    df = pd.DataFrame(
        [{"isin": e["isin"], "yield": e["yield"], "as_of": e["as_of"],
          "holders": "; ".join(e["holders"]), "source": e["source"],
          "mf_rating": e["mf_rating"], "instrument_name": e["name"]}
         for e in merged.values()])
    df = df.sort_values("isin")
    df.to_csv(out_csv, index=False)
    print(f"wrote {out_csv}: {len(df)} ISINs, {df['yield'].notna().sum()} with yield")
    return df


if __name__ == "__main__":
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    AS_OF = "2026-06-30"
    files = {
        "HDFC":          sorted((base / "hdfc").glob("*.xlsx")),
        "SBI":           [base / "sbi" / "sbi.xlsx"],
        "ICICI Pru":     sorted((base / "icici").glob("*.xlsx")),
        "Aditya Birla":  sorted((base / "absl").glob("*.xlsx")),
        "Axis":          [base / "axis" / "axis.xlsx"],
        "Nippon":        [base / "nippon" / "nippon.xls"],
        "Kotak":         [base / "kotak" / "kotak.xlsx"],
        "UTI":           sorted((base / "uti").glob("Sebi Exposure*.xls*")),
    }
    as_of = {k: AS_OF for k in files}
    build(files, as_of, base / "mf_yields.csv")
